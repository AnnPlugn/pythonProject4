import tkinter.messagebox
import tkinter as tk

import openpyxl
import pymysql.cursors
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

import DataBase
import pandas as pd


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.db_name = None
        self.table_name = None
        self.title("Создание базы данных")

        self.db_label = tk.Label(self, text="Имя базы данных:")
        self.db_label.pack()
        self.db_entry = tk.Entry(self, textvariable=tk.StringVar())
        self.db_entry.pack()

        self.table_label = tk.Label(self, text="Имя таблицы:")
        self.table_label.pack()
        self.table_entry = tk.Entry(self, textvariable=tk.StringVar())
        self.table_entry.pack()

        self.osn1 = tk.Label(self, text="Первое основание:")
        self.osn1.pack()
        self.osn1 = tk.Entry(self, textvariable=tk.StringVar())
        self.osn1.pack()

        self.osn2 = tk.Label(self, text="Второе основание:")
        self.osn2.pack()
        self.osn2 = tk.Entry(self, textvariable=tk.StringVar())
        self.osn2.pack()

        self.height = tk.Label(self, text="Высота:")
        self.height.pack()
        self.height = tk.Entry(self, textvariable=tk.StringVar())
        self.height.pack()

        self.file1_label1 = tk.Label(self, text="Имя файла эксель:")
        self.file1_label1.pack()
        self.file1_entry1 = tk.Entry(self, textvariable=tk.StringVar())
        self.file1_entry1.pack()

        self.db_label.pack(padx=50, pady=2)
        self.table_label.pack(padx=50, pady=2)
        self.osn1.pack(padx=50, pady=2)
        self.osn2.pack(padx=50, pady=2)
        self.height.pack(padx=50, pady=2)
        self.file1_label1.pack(padx=50, pady=2)

    def create_database(self):
        self.db_name = self.db_entry.get()
        self.table_name = self.table_entry.get()

        db = DataBase.DataBase(self.db_name, self.table_name)
        db.check_db()
        db.check_table()

    def save_result(self):
        osn1 = self.osn1.get()
        osn2 = self.osn2.get()
        height = self.height.get()
        osn1 = float(osn1)
        osn2 = float(osn2)
        height = float(height)
        trap = lambda a, b, c: (a + b) * c / 2
        res = trap(osn1,osn2,height)
        formatted_result = "{:.2f}".format(res)
        db1 = DataBase.DataBase(self.db_name, self.table_name)
        connection = db1.con_db()

        try:
            with connection.cursor() as cursor:
                cursor.execute(f"INSERT INTO {db1.name_tb} (osn1, osn2, height, S) VALUES (%s, %s, %s, %s)",
                               (osn1, osn2, height, formatted_result))
                connection.commit()

                cursor.execute(f"SELECT * FROM {db1.name_tb}")
                print(cursor.fetchall()[-1])

        except pymysql.err.DataError as e:
            print('Ошибка с данными:', e)

        except pymysql.err.DatabaseError as e:
            print(e)

    def list_tb(self):
        db1 = DataBase.DataBase(self.db_name, self.table_name)
        connection = db1.con_db()
        cursor = connection.cursor()
        tb_in_db = "SHOW TABLES;"
        cursor.execute(tb_in_db)
        tables = cursor.fetchall()

        table_list = [table[0] for table in tables]
        table_list_str = "\n".join(table_list)

        tkinter.messagebox.showinfo("Список таблиц", table_list_str)

    def save_to_excel(self):
        db1 = DataBase.DataBase(self.db_name, self.table_name)
        connection = db1.con_db()
        try:
            new_df = pd.read_sql("SELECT * FROM " + self.table_name, connection)
            wb = openpyxl.Workbook()
            ws = wb.active

            for r in dataframe_to_rows(new_df, index=False, header=True):
                ws.append(r)

            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except TypeError:
                        pass

                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
            file1 = self.file1_entry1.get()
            wb.save(file1)
            print(new_df)

            tkinter.messagebox.showinfo("Импорт в эксель", file1)

        except pymysql.err.DatabaseError as e:
            print(e)
        return


app = App()

create_button1 = tk.Button(app, text="Создать БД", command=app.create_database)
create_button1.pack()

create_button = tk.Button(app, text="Создать запись", command=app.save_result)
create_button.pack()

list_button = tk.Button(app, text="Показать список таблиц", command=app.list_tb)
list_button.pack()

excel_button = tk.Button(app, text="Импорт в эксель", command=app.save_to_excel)
excel_button.pack()

app.mainloop()
